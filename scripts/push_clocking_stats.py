#!/usr/bin/env python3
"""
Push clocking report summary stats to the workspace-dashboard GitHub repo.
Called by build_report.py after each successful report run.

Reads the just-generated Excel master, extracts key stats, writes
clocking_stats.json into the repo, then commits and pushes.

Usage:
    python push_clocking_stats.py --report <path/to/Clocking Report *.xlsx>

Environment:
    WORKSPACE_DASHBOARD_REPO  — absolute path to the workspace-dashboard repo
                                (defaults to C:\\Users\\quint\\workspace-dashboard)
"""

import argparse
import json
import os
import subprocess
import sys
from datetime import datetime

try:
    import pandas as pd
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: pandas / openpyxl not installed.")
    sys.exit(1)

REPO_DEFAULT = r"C:\Users\quint\workspace-dashboard"
OUTPUT_FILE  = "clocking_stats.json"

parser = argparse.ArgumentParser()
parser.add_argument("--report", required=True, help="Path to the Clocking Report xlsx")
args = parser.parse_args()

repo = os.environ.get("WORKSPACE_DASHBOARD_REPO", REPO_DEFAULT)
out  = os.path.join(repo, OUTPUT_FILE)

if not os.path.isfile(args.report):
    print(f"ERROR: Report not found: {args.report}")
    sys.exit(1)

# ── Extract stats from the Clocking Report sheet ──────────────────────────────
print(f"Reading {os.path.basename(args.report)}...")

try:
    # Row 2 is the subtitle (contains period, employee count, records)
    wb   = load_workbook(args.report, read_only=True, data_only=True)
    ws   = wb["Clocking Report"]
    subtitle = ws["A2"].value or ""

    df_detail = pd.read_excel(args.report, sheet_name="Clocking Report", header=3)
    df_emp    = pd.read_excel(args.report, sheet_name="Summary by Employer", header=3)
    df_miss   = pd.read_excel(args.report, sheet_name="Missing Clock Out",   header=3)
    wb.close()
except Exception as e:
    print(f"ERROR reading report: {e}")
    sys.exit(1)

# Derive period from subtitle ("Period: DD Mon – DD Mon YYYY")
period = ""
if "Period:" in subtitle:
    period = subtitle.split("Period:")[1].split("|")[0].strip()

n_employees = int(df_detail["Employee ID"].nunique()) if "Employee ID" in df_detail.columns else 0
n_records   = len(df_detail)
n_missing   = len(df_miss)

# Per-employer breakdown
employer_stats: dict = {}
for _, row in df_emp.iterrows():
    name = str(row.get("Employer", "")).strip()
    if not name or name.upper() in ("EMPLOYER", "TOTAL"):
        continue
    employer_stats[name] = {
        "employees":      int(row.get("Unique Employees", 0) or 0),
        "total_hours":    str(row.get("Total Hours", "") or ""),
        "missing_clkout": int(row.get("Missing Clock Out", 0) or 0),
    }

stats = {
    "generated_at":     datetime.now().strftime("%Y-%m-%d %H:%M SAST"),
    "report_file":      os.path.basename(args.report),
    "period":           period,
    "total_employees":  n_employees,
    "total_records":    n_records,
    "missing_clockout": n_missing,
    "by_employer":      employer_stats,
}

# ── Write JSON ────────────────────────────────────────────────────────────────
with open(out, "w", encoding="utf-8") as f:
    json.dump(stats, f, indent=2)
print(f"Written: {out}")

# ── Git commit + push ─────────────────────────────────────────────────────────
def git(cmd: list[str]) -> tuple[int, str]:
    r = subprocess.run(cmd, cwd=repo, capture_output=True, text=True)
    return r.returncode, (r.stdout + r.stderr).strip()

code, out_txt = git(["git", "add", OUTPUT_FILE])
if code:
    print(f"ERROR: git add failed:\n{out_txt}")
    sys.exit(1)

code, out_txt = git(["git", "diff", "--staged", "--quiet"])
if code == 0:
    print("No changes to clocking_stats.json — skipping commit.")
    sys.exit(0)

today   = datetime.now().strftime("%Y-%m-%d")
msg     = f"chore: update clocking stats {today}"
code, out_txt = git(["git", "commit", "-m", msg])
if code:
    print(f"ERROR: git commit failed:\n{out_txt}")
    sys.exit(1)
print(f"Committed: {msg}")

code, out_txt = git(["git", "push"])
if code:
    print(f"ERROR: git push failed:\n{out_txt}")
    sys.exit(1)
print("Pushed to GitHub.")
print(f"  Period:    {stats['period']}")
print(f"  Employees: {stats['total_employees']}")
print(f"  Records:   {stats['total_records']}")
print(f"  Missing:   {stats['missing_clockout']}")
